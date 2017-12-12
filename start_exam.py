import config
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.keys import Keys
from pyquery import PyQuery as pq
from openpyxl import load_workbook
from time import sleep


def login():
    driver =config.DRIVER
    driver.get(config.LOGIN_URL)
    input("完成登陆后按回车键继续")

    try:
        WebDriverWait(driver,config.TIMEOUT).until(EC.presence_of_element_located((By.CSS_SELECTOR,"div.indexwdks a[href]")))
    except TimeoutException:
        print("超时")

    try:
        exam_btn = driver.find_element_by_css_selector("div.indexwdks a[href]")
        exam_btn.send_keys(Keys.ENTER)
    except NoSuchElementException:
        print("没有找到我的考试")

    try:
        WebDriverWait(driver, config.TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "ul#currcontainer a[href]")))
    except TimeoutException:
        print("超时")

    try:
        exam_btn = driver.find_element_by_css_selector("ul#currcontainer a[href]")
        exam_btn.send_keys(Keys.ENTER)
    except NoSuchElementException:
        print("没有找到进入考试 ")

    switch_new(driver)

    try:
        WebDriverWait(driver, config.TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "div#timucontent")))
    except TimeoutException:
        print("超时")
    return driver



def answer(driver):
    # 检索
    wb = load_workbook(filename=config.ANSWER_PATH)
    one = wb.get_sheet_by_name('one')
    multi = wb.get_sheet_by_name('multi')
    true = wb.get_sheet_by_name('true')
    while(1):
        doc = get_html(driver)
        ques = doc('#timucontent h2').text().replace(' ', '')
        find_answer(driver,ques,one,multi,true)
        sleep(0.2)
        try:
            next = driver.find_element_by_css_selector('[id=nextButton]')
            next.click()
        except NoSuchElementException:
            print("没有找到下一题 ")
            break

def find_answer(driver,ques,one,multi,true):
    # 题号
    page = ques.split('、')[0]
    # 类别
    cat = ques.split('、')[1][:5]
    # 关键词
    ques_list = ques.split('）')
    keyword = sorted(ques_list, key=lambda x: len(x))[-1]
    # print("keyword =",keyword,'\n类别=',cat)
    if(cat=='（单选题）'):
        sel=''
        for row in range(1, one.max_row + 1):
            if (keyword in one.cell(row=row, column=1).value):
                # print(page, one.cell(row=row, column=1).value, '\n', one.cell(row=row, column=2).value, '\n')
                sel= one.cell(row=row, column=3).value
                break
        try:
            option = driver.find_element_by_css_selector('input[value='+sel+']')
            option.click()
        except NoSuchElementException:
            print("未答！：\n",ques)
    if(cat=='（多选题）'):
        sels=''
        for row in range(1, multi.max_row + 1):
            if (keyword in multi.cell(row=row, column=1).value):
                # print(page, multi.cell(row=row, column=1).value, '\n', multi.cell(row=row, column=2).value, '\n')
                sels= multi.cell(row=row, column=3).value
                break
        try:
            # print(page,sels)
            for sel in sels:
                option = driver.find_element_by_css_selector('input[value='+sel+']')
                option.click()
        except NoSuchElementException:
            print("未答！：\n", ques)
    if(cat=='（判断题）'):
        judge=''
        for row in range(1, true.max_row + 1):
            if (keyword in true.cell(row=row, column=1).value):
                # print(page, true.cell(row=row, column=1).value)
                judge= true.cell(row=row, column=2).value
                break
        try:
            option = driver.find_element_by_css_selector('input[value=\"'+str(judge)+'\"]')
            option.click()
        except NoSuchElementException:
            print("未答！：\n", ques)

def get_html(driver):
    html = driver.page_source
    # 必须去掉xmlns属性，不然不能正确解析
    html = html.replace('xmlns','another_attr')
    return pq(html)

def switch_new(driver):
    # 获取所有页面的句柄，并循环判断不是当前的句柄
    for handle in driver.window_handles:
        if(handle != driver.current_window_handle):
            driver.close()
            driver.switch_to.window(handle)
            break

if __name__ =='__main__':
    driver = login()
    answer(driver)